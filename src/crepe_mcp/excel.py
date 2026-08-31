"""Excel (.xlsx) file creation, inspection, and manipulation tools for CREPE MCP server.

Uses openpyxl for fast, pure-Python spreadsheet creation, formatting, and reading.
"""
from __future__ import annotations

import os
import re
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_excel(
    output_path: str,
    sheets: list[dict[str, Any]] | None = None,
    overwrite: bool = True,
) -> dict:
    """Create a new Excel (.xlsx) workbook with styled header rows and data.

    output_path : absolute path for the generated file.
    sheets      : list of dicts:
        [
            {
                "name": "Sheet1",
                "headers": ["Name", "Score", "Notes"],
                "rows": [["Alice", 95, "Pass"], ["Bob", 88, "Pass"]]
            }
        ]
    """
    if not os.path.isabs(output_path):
        return {"success": False, "error": f"output_path must be an absolute path, got {output_path!r}"}

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    if os.path.exists(output_path) and not overwrite:
        return {"success": False, "error": f"File already exists: {output_path!r}"}

    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(wb.active)


    sheet_list = sheets if (sheets is not None and len(sheets) > 0) else [{"name": "Sheet1", "headers": [], "rows": []}]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    created_sheets = []
    for s_info in sheet_list:
        s_name = s_info.get("name", "Sheet1")
        ws = wb.create_sheet(title=s_name)

        headers = s_info.get("headers", [])
        rows = s_info.get("rows", [])

        if headers:
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in rows:
            ws.append(row)

        # Style data cells and auto-adjust column width
        for row in ws.iter_rows(min_row=2 if headers else 1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if val_str:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        created_sheets.append(s_name)

    try:
        wb.save(output_path)
    except Exception as exc:
        return {"success": False, "error": f"Failed to save Excel file: {exc}"}

    return {
        "success": True,
        "output_path": output_path,
        "sheets": created_sheets,
        "size_bytes": os.path.getsize(output_path),
    }


def inspect_excel(input_path: str, max_rows: int = 100) -> dict:
    """Inspect an existing .xlsx file and return sheets, dimensions, cell values, and formulas."""
    if not os.path.isabs(input_path):
        return {"success": False, "error": f"input_path must be an absolute path, got {input_path!r}"}
    if not os.path.isfile(input_path):
        return {"success": False, "error": f"File not found: {input_path!r}"}

    try:
        wb = openpyxl.load_workbook(input_path, data_only=False)
    except Exception as exc:
        return {"success": False, "error": f"Failed to load workbook: {exc}"}

    sheet_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_preview = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx > max_rows:
                break
            rows_preview.append([str(c) if c is not None else "" for c in row])

        sheet_data.append({
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "preview_rows": len(rows_preview),
            "rows": rows_preview,
        })

    return {
        "success": True,
        "input_path": input_path,
        "sheet_count": len(sheet_data),
        "sheets": sheet_data,
    }


def update_excel_sheet(
    input_path: str,
    sheet_name: str,
    append_rows: list[list[Any]] | None = None,
    update_cells: dict[str, Any] | None = None,
) -> dict:
    """Update an existing .xlsx sheet by appending rows or setting cell values
    (e.g. update_cells={"B2": 42, "C2": "=SUM(A1:B1)"}).
    """

    if not os.path.isabs(input_path):
        return {"success": False, "error": f"input_path must be an absolute path, got {input_path!r}"}
    if not os.path.isfile(input_path):
        return {"success": False, "error": f"File not found: {input_path!r}"}

    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as exc:
        return {"success": False, "error": f"Failed to load workbook: {exc}"}

    if sheet_name not in wb.sheetnames:
        return {"success": False, "error": f"Sheet {sheet_name!r} not found in workbook. Available: {wb.sheetnames}"}

    ws = wb[sheet_name]

    try:
        if update_cells:
            for cell_coord, val in update_cells.items():
                ws[cell_coord] = val

        if append_rows:
            for r in append_rows:
                ws.append(r)

        wb.save(input_path)
    except Exception as exc:
        return {"success": False, "error": f"Failed to update workbook: {exc}"}

    return {
        "success": True,
        "input_path": input_path,
        "sheet_name": sheet_name,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }


def markdown_table_to_excel(
    markdown_table: str,
    output_path: str,
    sheet_name: str = "Data",
) -> dict:
    """Convert a GitHub Flavored Markdown table string into an Excel workbook."""
    lines = [line.strip() for line in markdown_table.strip().splitlines() if line.strip()]
    if not lines:
        return {"success": False, "error": "Empty Markdown table input"}

    parsed_rows = []
    for line in lines:
        # Ignore delimiter rows like |---|---|
        if re.match(r"^\|?\s*:?-+:?\s*(\|?\s*:?-+:?\s*)+\|?$", line):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        parsed_rows.append(cells)

    if not parsed_rows:
        return {"success": False, "error": "No valid table rows found"}

    headers = parsed_rows[0]
    data_rows = parsed_rows[1:]

    return create_excel(
        output_path=output_path,
        sheets=[{"name": sheet_name, "headers": headers, "rows": data_rows}],
        overwrite=True,
    )
